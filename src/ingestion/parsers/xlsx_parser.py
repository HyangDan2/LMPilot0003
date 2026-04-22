from __future__ import annotations

from pathlib import Path
from typing import Any

from src.ingestion.parsers.base import DocumentParser, ParserError
from src.models.schemas import ParsedDocument, Section


class XlsxParser(DocumentParser):
    """MVP XLSX parser that converts sheets into capped TSV-like sections."""

    file_type = "xlsx"
    max_rows = 100
    max_columns = 30

    def parse(self, path: Path) -> ParsedDocument:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except Exception as exc:
            raise ParserError("XLSX parsing requires openpyxl.") from exc

        try:
            workbook = load_workbook(str(path), data_only=True, read_only=True)
        except Exception as exc:
            raise ParserError(f"Failed to open XLSX {path}: {exc}") from exc

        doc_id = self.doc_id(path)
        sections: list[Section] = []
        sheet_texts: list[str] = []
        try:
            for index, sheet in enumerate(workbook.worksheets, start=1):
                text = self._sheet_to_tsv(sheet)
                if text:
                    sheet_texts.append(f"Sheet: {sheet.title}\n{text}")
                sections.append(
                    Section(
                        section_id=f"{doc_id}-sheet-{index}",
                        title=sheet.title,
                        level=1,
                        text=text,
                        page_or_slide=sheet.title,
                        metadata={
                            "parser": "xlsx",
                            "max_rows": self.max_rows,
                            "max_columns": self.max_columns,
                            "source_max_row": sheet.max_row,
                            "source_max_column": sheet.max_column,
                        },
                    )
                )
        finally:
            workbook.close()

        return ParsedDocument(
            doc_id=doc_id,
            file_path=str(path.resolve()),
            file_type=self.file_type,
            title=self.title_from_path(path),
            text="\n\n".join(sheet_texts).strip(),
            sections=sections,
            assets=[],
            metadata={"sheet_count": len(sections)},
        )

    def _sheet_to_tsv(self, sheet: Any) -> str:
        lines: list[str] = []
        for row in sheet.iter_rows(max_row=self.max_rows, max_col=self.max_columns, values_only=True):
            values = ["" if value is None else str(value).replace("\n", " ").strip() for value in row]
            if any(values):
                lines.append("\t".join(values).rstrip())
        return "\n".join(lines).strip()

